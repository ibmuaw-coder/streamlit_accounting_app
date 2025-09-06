# streamlit_accounting_app.py
# A Streamlit-based accounting app that creates/uses an Excel file (accounting_data.xlsx),
# initializes required sheets and a chart of accounts (with sample accounts requested),
# allows adding journal entries, basic sales entries, OCR/text input, and uses OpenAI API
# to analyze / suggest structure for free-text inputs.
#
# Requirements (put in requirements.txt):
# streamlit
# pandas
# openpyxl
# openai
# pillow
# pytesseract (optional, for OCR)
# opencv-python (optional, for image preprocessing)
# python-dotenv (optional)
#
# Notes:
# - Do NOT commit your OPENAI_API_KEY to GitHub. Use Streamlit secrets or environment vars.
# - To deploy on Streamlit Community Cloud: push this repository to GitHub, then in
#   https://streamlit.io/ deploy the app from your repo and set the secret OPENAI_API_KEY.
# - This app is a starting template and intentionally keeps heavy features optional
#   (OCR, audio) because they may require system-level binaries on the server.

import streamlit as st
import pandas as pd
import openpyxl
import openai
import io
import os
import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PIL import Image

# Optional imports for OCR/image processing. If not available, features will be disabled gracefully.
try:
    import pytesseract
    import cv2
    OCR_AVAILABLE = True
except Exception:
    OCR_AVAILABLE = False

# --------------------------- Configuration ---------------------------------
EXCEL_FILENAME = "accounting_data.xlsx"
SHEETS = [
    'اليومية', 'المبيعات', 'المشتريات', 'العملاء', 'الموردين', 'المخزون',
    'قائمة الدخل', 'الميزانية', 'قائمة الحسابات'
]

# Chart of accounts: requested accounts and some common ones
DEFAULT_ACCOUNTS = [
    (1000, 'الأصول'),
    (1100, 'الأصول المتداولة'),
    (1110, 'الصندوق'),
    (1120, 'البنك'),
    (1130, 'الذمم المدينة'),
    (1200, 'الأصول الثابتة'),
    (2000, 'الخصوم'),
    (2100, 'الذمم الدائنة'),
    (3000, 'حقوق الملكية'),
    (4000, 'المبيعات'),
    (5000, 'تكلفة المبيعات'),
    (6000, 'المصروفات'),
]

# --------------------------- Utility functions ----------------------------

def get_openai_api_key():
    """Retrieve OpenAI API key from Streamlit secrets, env var, or user input."""
    # 1) Streamlit secrets
    if "OPENAI_API_KEY" in st.secrets:
        return st.secrets["OPENAI_API_KEY"]
    # 2) Environment variable
    if os.getenv("OPENAI_API_KEY"):
        return os.getenv("OPENAI_API_KEY")
    # 3) Session state (entered by user during this session)
    return st.session_state.get("openai_api_key", None)


def set_openai_key_in_session(key: str):
    st.session_state["openai_api_key"] = key


def init_workbook_if_missing(filename=EXCEL_FILENAME):
    """Create the Excel workbook and required sheets with headers if missing."""
    if not os.path.exists(filename):
        wb = Workbook()
        # remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        for name in SHEETS:
            wb.create_sheet(name)

        # Initialize headers
        ws = wb['اليومية']
        ws.append(["التاريخ", "البيان", "الحساب", "مدين", "دائن", "الرصيد"])

        ws = wb['المبيعات']
        ws.append(["رقم الفاتورة", "التاريخ", "العميل", "المبلغ", "الضريبة", "الإجمالي"])

        ws = wb['المشتريات']
        ws.append(["رقم الفاتورة", "التاريخ", "المورد", "المبلغ", "الضريبة", "الإجمالي"])

        ws = wb['قائمة الحسابات']
        ws.append(["رقم الحساب", "اسم الحساب", "نوع الحساب"])
        for acc_num, acc_name in DEFAULT_ACCOUNTS:
            acc_type = 'مدين' if str(acc_num).startswith('1') else 'دائن' if str(acc_num).startswith('2') else 'ملكي/دخل/مصروف'
            ws.append([acc_num, acc_name, acc_type])

        wb.save(filename)


def load_sheet_as_df(filename, sheet_name):
    try:
        df = pd.read_excel(filename, sheet_name=sheet_name)
        return df
    except Exception:
        return pd.DataFrame()


def append_row_to_sheet(filename, sheet_name, row_values):
    wb = openpyxl.load_workbook(filename)
    ws = wb[sheet_name]
    ws.append(row_values)
    wb.save(filename)


# --------------------------- OpenAI integration ---------------------------

def analyze_with_openai(text: str, api_key: str):
    """Send a prompt to OpenAI and return a JSON-friendly response.
    This uses the Chat Completions (ChatGPT) API via the openai package.
    """
    if not api_key:
        raise RuntimeError("OpenAI API key is missing. Set OPENAI_API_KEY in Streamlit secrets or paste it in the sidebar.")

    openai.api_key = api_key

    # Build a helpful system + user prompt in Arabic to parse simple accounting text
    system = {
        "role": "system",
        "content": "أنت مساعد محاسبة ذكي. أعدّل واطّلع على المدخلات النصية المحاسبية وأعدَّ مخرجات منظمة بصيغة JSON بدون شرح إضافي.\n" 
                   "المخرجات المتوقعة: {type: 'invoice'|'journal'|'other', date: 'YYYY-MM-DD', lines: [{account: 'اسم الحساب', debit: 0.0, credit: 0.0, desc: ''}], total: 0.0, currency: 'USD' }"
    }

    user_msg = {"role": "user", "content": f"حلل النص التالي وأعطه شكل قيد محاسبي JSON:\n{text}"}

    # Use chat completion. Model name can be changed if you have access to others.
    try:
        resp = openai.ChatCompletion.create(
            model="gpt-4o",
            messages=[system, user_msg],
            temperature=0.2,
            max_tokens=800
        )
        content = resp['choices'][0]['message']['content']
        return content
    except Exception as e:
        return f"ERROR: {str(e)}"


# --------------------------- Streamlit UI --------------------------------

st.set_page_config(page_title="نظام المحاسبة الذكي - Excel AI", layout="wide")

st.sidebar.title("إعدادات")
api_key_input = st.sidebar.text_input("OpenAI API Key (أدخل أو ضع في Secrets)", type="password")
if api_key_input:
    set_openai_key_in_session(api_key_input)

st.sidebar.markdown("**ملاحظات أمان**: لا تقم بحفظ المفتاح في الكود أو رفعه إلى GitHub. استخدم Streamlit Secrets أو متغير بيئة.")

st.title("نظام المحاسبة الذكي - Excel AI")
col1, col2 = st.columns([2,1])

with col2:
    st.write("الوقت الحالي:")
    st.write(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

with col1:
    st.write("مرحباً! استخدم الواجهات لادخال البيانات أو رفع ملفات أو طلب تحليل نصي بواسطة GPT.")

# Initialize workbook if missing
if st.button("تهيئة/إنشاء ملف Excel الافتراضي"):
    init_workbook_if_missing()
    st.success(f"تم إنشاء أو التأكد من وجود الملف: {EXCEL_FILENAME}")

# Show Chart of Accounts
st.subheader("قائمة الحسابات (دفتر الحسابات)")
if st.button("عرض قائمة الحسابات"):
    df_accounts = load_sheet_as_df(EXCEL_FILENAME, 'قائمة الحسابات')
    st.dataframe(df_accounts)

# Section: Manual journal entry
st.subheader("إضافة قيد يدوي")
with st.form(key='journal_form'):
    j_date = st.date_input("التاريخ", value=datetime.date.today())
    j_desc = st.text_input("البيان")

    # load accounts for selector
    accounts_df = load_sheet_as_df(EXCEL_FILENAME, 'قائمة الحسابات')
    account_names = accounts_df['اسم الحساب'].tolist() if not accounts_df.empty else [a[1] for a in DEFAULT_ACCOUNTS]
    j_account = st.selectbox("الحساب", options=account_names)
    colA, colB = st.columns(2)
    with colA:
        j_debit = st.number_input("مدين", min_value=0.0, value=0.0)
    with colB:
        j_credit = st.number_input("دائن", min_value=0.0, value=0.0)

    submit = st.form_submit_button("معاينة وإرسال إلى Excel (بعد موافقة GPT إذا طلبت)")

if submit:
    # show preview
    preview = {
        'date': str(j_date), 'desc': j_desc, 'account': j_account,
        'debit': j_debit, 'credit': j_credit
    }
    st.write("معاينة القيد:")
    st.json(preview)

    # Optional analysis with OpenAI
    api_key = get_openai_api_key()
    if api_key:
        st.info("إرسال النص للتحقق والتحليل من قبل GPT...")
        text_to_analyze = f"قيد: التاريخ {j_date}, البيان {j_desc}, الحساب {j_account}, مدين {j_debit}, دائن {j_credit}"
        analysis = analyze_with_openai(text_to_analyze, api_key)
        st.write("تحليل GPT:")
        st.text(analysis)
    else:
        st.warning("لم يتم العثور على مفتاح OpenAI. يمكنك وضعه في الشريط الجانبي أو في Streamlit secrets.")

    # Commit to excel
    if st.button("حفظ القيد في Excel"):
        try:
            append_row_to_sheet(EXCEL_FILENAME, 'اليومية', [str(j_date), j_desc, j_account, float(j_debit), float(j_credit), ""])
            st.success("تم حفظ القيد في الورقة 'اليومية'.")
        except Exception as e:
            st.error(f"حدث خطأ أثناء حفظ القيد: {str(e)}")

# Section: Text input processed by GPT
st.subheader("تحليل نصي تلقائي (أدخل وصف فاتورة أو سند)")
user_text = st.text_area("أدخل النص هنا:", height=150)
if st.button("تحليل النص بواسطة GPT"):
    api_key = get_openai_api_key()
    if not api_key:
        st.error("مفتاح OpenAI مفقود. أدخله في الشريط الجانبي أو في Streamlit secrets.")
    else:
        result = analyze_with_openai(user_text, api_key)
        st.write("نتيجة التحليل:")
        st.text(result)

# Section: OCR from Image (optional)
st.subheader("رفع صورة لتحويل النص (OCR) - اختياري")
img_file = st.file_uploader("اختر صورة (jpg/png)", type=['jpg','jpeg','png'])
if img_file is not None:
    st.image(Image.open(img_file), caption='الصورة المرفوعة', use_column_width=True)
    if OCR_AVAILABLE:
        try:
            img_bytes = img_file.read()
            img = Image.open(io.BytesIO(img_bytes)).convert('RGB')
            text = pytesseract.image_to_string(img, lang='ara+eng')
            st.text_area("النص المستخرج:", text, height=200)
            if st.button("تحليل النص المستخرج بواسطة GPT"):
                api_key = get_openai_api_key()
                if not api_key:
                    st.error("مفتاح OpenAI مفقود.")
                else:
                    analysis = analyze_with_openai(text, api_key)
                    st.text(analysis)
        except Exception as e:
            st.error(f"فشل استخراج النص: {str(e)}")
    else:
        st.warning("خاصية OCR غير متاحة هنا لأن pytesseract/Opencv غير مثبتان في البيئة. يمكنك تفعيلها محلياً.")

# Section: Download current Excel
st.subheader("تنزيل ملف Excel الحالي")
if os.path.exists(EXCEL_FILENAME):
    with open(EXCEL_FILENAME, 'rb') as f:
        data = f.read()
    st.download_button(label='تحميل ملف Excel', data=data, file_name=EXCEL_FILENAME)
else:
    st.info("لم يتم العثور على ملف Excel. اضغط على زر التهيئة لإنشائه.")

# Footer: short help
st.markdown("---")
st.markdown("**تعليمات النشر على Streamlit:** ادخل مفتاح OpenAI في Streamlit Secrets (Settings -> Secrets) باسم OPENAI_API_KEY، ثم أنشئ التطبيق من مستودع GitHub. راجع README في المستودع للمزيد.")

# End of file
